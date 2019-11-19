# -*- coding: utf-8 -*-

# ログ確認用のモジュール
import logging
logging.basicConfig(format='%(levelname)s:%(message)s', level=logging.INFO)

# OS依存の機能を扱うモジュール
import os
# TwitterAPIを扱うモジュール
import tweepy
# EXCELファイルをPythonで扱うモジュール
import openpyxl as op
# 正規表現を扱うモジュール
import re
#Flaskを扱うモジュール
from flask import Flask, request, redirect, render_template, make_response, session
# ファイル名をチェックするモジュール
from werkzeug.utils import secure_filename

# カラープリセット
color = {0:"ff7f7f", 1:"ff7fbf", 2:"ff7fff", 3:"bf7fff", 4:"7f7fff",
         5:"7fbfff", 6:"7fffff", 7:"7fffbf", 8:"7fff7f", 9:"bfff7f",
         10:"ffff7f", 11:"ffbf7f", 12:"ff0000", 13:"ff007f", 14:"ff00ff",
         15:"007fff", 16:"00ffff", 17:"00ff7f", 18:"00ff00", 19:"7fff00",
         20:"ffff00", 21:"ff7f00", 22:"fcc800", 23:"9cbb1c", 24:"00a960"}
# 色参考https://www.colordic.org/p/

# 抽出元ツイートインスタンス
class Source_Tweet():
    def __init__(self, user_screen_name, user_name, text, chara, source_url, description):
        # ユーザーID、ユーザー名、ツイート本文テキスト、キャラクター名、抽出元URL、プロフィールテキスト
        self.user_screen_name
        self.user_name
        self.text
        self.chara
        self.source_url
        self.description

# サークルインスタンス
class Circle():
    def __init__(self, num, user_name, user_screen_name, circle_name, chara, source_url):
        # スペース番号、ユーザー名、ユーザーID、サークル名、キャラクター名、抽出元URL
        self.num
        self.user_name
        self.user_screen_name
        self.chara
        self.source_url
    
    def return_list(self):
        # 全ての要素をリストにして返す
        return [self.num,
                self.user_name,
                self.user_screen_name,
                self.chara,
                self.source_url]

#*--------初期設定--------*
# Consumer Key
CONSUMER_KEY = os.environ["CONSUMER_KEY"]
# Consumer Secret
CONSUMER_SECRET =os.environ["CONSUMER_SECRET"]
# Callback URL (認証後リダイレクトされるURL)
#CALLBACK_URL = 'https://circle-check-app.herokuapp.com/' # Heroku上
CALLBACK_URL = 'https://circle-check-app.azurewebsites.net' # azure上
# ファイルをダウンロードした際、クライアント側で適切にファイルを処理できるようにmimetypeを定義
# mimetype:ファイル形式をサーバーに認識させるための識別子
# 参考:https://docs.microsoft.com/ja-jp/previous-versions/office/office-2007-resource-kit/ee309278(v=office.12)
XLSX_MIMETYPE = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
# ファイルのアップロード先のディレクトリ
UPLOAD_FOLDER = './uploads'
# テンプレート用ファイルのディレクトリ
DEFAULT_FOLDER = './defaults'
# アップロードされる拡張子の制限
ALLOWED_EXTENSIONS = set(['xlsx'])
# flaskの起動
app = Flask(__name__)
# flask の session を使うにはkeyを設定する必要がある．
app.secret_key = os.environ["SECRET_KEY"]
# フォルダディレクトリを保存
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['DEFAULT_FOLDER'] = DEFAULT_FOLDER 

#-------- rootページ --------
@app.route('/', methods=['GET', 'POST']) # GETとPOSTのみ
def index(): # rootページ読み込み時にindex()を実行する
    if request.method == "GET": # リクエストメソッドがGETのとき(twitter認証関連)
        api = get_api()                                 # Twitter_APIの取得
        if api:
            logging.info("API認証済み")
        else:
            logging.info("API未認証")
        return render_template("index.html", api=api)   # indexページの表示
    
    elif request.method == "POST": #リクエストメソッドがPOSTのとき(ファイル受け渡し関連)
        try:
            api = app.config['API'] # APiオブジェクトの取得
            logging.info("API......OK!")
        except: # APIオブジェクトがない
            logging.warning("API......lost?")
            return render_template("oauth_error.html", error = 1)
        
        logging.info("---- Start ----")
        filename = upload() # アップされたファイルの読み込み
        logging.info("Event:" + request.form["events"])
        logging.info("Character:" + request.form["character"])
        
        try:
            if filename:                #ファイル名が存在すれば
                #ツイート格納用
                tweet = []
                
                #検索条件の指定
                if (request.form["events"] == "" or request.form["character"] == ""):
                    return render_template("index.html", api=app.config['API'], error=3)
                    
                event = "(" + request.form["events"].replace("　"," ").replace(" "," OR ") + ")"     #イベント名入力(ORで接続)
                character = request.form["character"].replace("　"," ").strip(" ").split(" ")        #キャラ名入力(全角空白を半角にし、両端の空白を削除、空白で区切りリスト化)
                if request.form["etc"]:# 日を跨ぐ即売会のとき
                    etc = "(" + request.form["etc"].replace("　"," ").replace(" "," OR ") + ")"          #曜日・日にち入力(日を跨ぐ即売会用)
                else: # 日を跨ぐ即売会でないとき
                    etc = ""
                    
                for chara in character:
                    query = event + " " + chara + " 新刊 -RT -filter:replies " + etc  #検索文字列群生成
                
                    logging.info("Query:" + query)
                
                    tweet_id = [] #ツイートID格納用
                    #ツイート情報の取得
                    try:
                        for status in api.search(q=query, lang='ja', result_type='recent', count=100, tweet_mode='extended'): 
                            #q:検索ワード("-RT"をつけることでRTを省ける) count:取得件数　lang:言語(日本語なら"ja") result_type:取得するツイート (recent時系列で取得) 
                            text = re.sub(r"(https?|ftp)(:\/\/[-_\.!~*\'()a-zA-Z0-9;\/?:\@&=\+\$,%#]+)", "" ,status.full_text) #URL部分を削除
                            #tweet:[ユーザーID, ユーザー名, ツイート本文, キャラクター名, ツイートURL, プロフ]
                            tweet.append(Source_Tweet(status.user.screen_name, 
                                          status.user.name, 
                                          text, 
                                          chara, 
                                          "https://twitter.com/"+ status.user.screen_name.strip("@") + "/status/" + str(status.id),
                                          status.user.description))
                            tweet_id.append(status.id) #ツイートIDの取得 api.search.id
                            #ユーザーごとに2次元配列で格納　.user.screen_name:UserID .user.name:Username .text:Tweet user.description:プロフィール
                        
                        if len(tweet_id) == 0:  #もしツイートがなかったら(これでmax_idを指定するとエラーになる)
                            continue            #次のキャラクターへ
                            
                        else:                   #ツイートがあれば追加で検索
                            #ツイート情報の追加取得(前のツイート取得の最後のmax_idより後ろのツイートを取得)
                            for i in range(9):
                                for status in api.search(q=query, lang='ja', result_type='recent', count=100, tweet_mode='extended', max_id=tweet_id[-1]-1): 
                                    #max_id - 指定されたID以下の（つまり、古い）IDを持つステータスのみを返す
                                    #リスト名[-1] でリストの一番最後の要素を取得 
                                    text = re.sub(r"(https?|ftp)(:\/\/[-_\.!~*\'()a-zA-Z0-9;\/?:\@&=\+\$,%#]+)", "" ,status.full_text) #URL部分を削除
                                    tweet.append(Source_Tweet(status.user.screen_name,
                                                              status.user.name, 
                                                              text,
                                                              chara,
                                                              "https://twitter.com/"+ status.user.screen_name.strip("@") + "/status/" + str(status.id),
                                                              status.user.description))
                                    tweet_id.append(status.id)
                                    
                    #API制限時処理
                    except tweepy.TweepError:
                        return render_template("index.html", api=app.config['API'], error=4)
                
                
                #正規表現のコンパイル
                num_pattern = re.compile("[a-zA-Zぁ-んァ-ヶ]-?[0-9]{2}[ab]?")             #スペース番号正規表現コンパイル
                circle_pattern = re.compile("サークル名?[「【『：:][\w\W]+?[」】』\n]")         #サークル名正規表現コンパイル
                
                #番号格納用
                No = []
                #サークル番号, ユーザー名, ユーザーID, サークル名, キャラクター名, 抽出元URL
                
                #パターンマッチング
                No = pattern_match(tweet, num_pattern, circle_pattern, No)
                
                #エクセル上での処理
                #ワークブックの読み込み
                wb = op.load_workbook(os.path.join(app.config['UPLOAD_FOLDER'], filename))  #xlsxファイルを開く
                
                #ワークシートの読み込み
                ws1 = wb.worksheets[0]
                ws2 = wb.worksheets[1]
                
                #マップ上にないスペース番号のサークル情報を削除する
                space_position = {} #マップ上のスペース番号の行列番号
                
                #マップ上のスペース番号の取得(同時にに罫線も設定する)
                try:
                    #罫線のフォーマット(黒の細線)
                    border = op.styles.borders.Border(top=op.styles.borders.Side(style='thin', color='000000'), 
                                      bottom=op.styles.borders.Side(style='thin', color='000000'), 
                                      left=op.styles.borders.Side(style='thin', color='000000'),
                                      right=op.styles.borders.Side(style='thin', color='000000'))
                    for ws1_row in ws1:
                        for ws1_cell in ws1_row:
                            if ws1_cell.value == None:  #セルの値がNoneのとき
                                continue
                            else:                       #セルの値が存在するとき
                                ws1_cell.border = border
                                space_position[ws1_cell.value] = [ws1_cell.row, ws1_cell.column] #{"スペース番号":[行番号, 列番号]}
                except: #マップに何も入力されていないとき
                    return render_template("index.html", api=app.config['API'], error = 5)
                
                No_list = [] #修正後のサークルリスト
                #マップ上のスペース番号のみを残す
                for num in No:
                    if re.sub("[ab]", "", num.num) in space_position: #マップ上に番号が存在したら
                        No_list.append(num)
                
                #サークルリストをシートに追加
                for i, row in enumerate(No_list):
                    ws2.append(row.return_list())                                                             #サークルリストの追加
                    ws2.cell(row = i+2, column = 3).hyperlink = "https://twitter.com/" + row.user_screen_name #ハイパーリンクの設定(ユーザーURL)
                    ws2.cell(row = i+2, column = 6).hyperlink = row.source_url                          #ハイパーリンクの設定(抽出元URL)
                        
                    
                #シートの書式設定
                sheet_format(ws2)
                
                #キャラクター一覧の取得
                #キャラクター一覧とスペース番号を紐付けする
                chara_dict = chara_set(ws2)[0]
                space_list = chara_set(ws2)[1]
                    
                #色付けをする
                coloring(space_list, chara_dict, space_position, ws1, ws2)                                             
                
                #ファイルの保存
                wb.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))    #xlsxファイルを保存する 
                
                logging.info("---- End ----")
                response = download(filename)   #ダウンロード用のレスポンスの作成
                return response                 #ダウンロード用のファイルのダウンロードタブ表示
        
            else:   #ファイル名が存在しないまたはxlsx形式でないとき
                return render_template("index.html", api=api, error = 1)
            
        except op.utils.exceptions.InvalidFileException:    #ファイル名に2バイト文字が含まれるとき
            return render_template("index.html", api=api, error = 2)  

#-------- 認証用ページ --------
@app.route('/twitter_auth', methods=['GET']) # GET以外のリクエストを拒否
def twitter_auth(): # 認証
    # tweepy でアプリのOAuth認証を行う
    logging.info("---- API認証開始 ----")
    #OAuthHandlerインスタンスの作成
    auth = tweepy.OAuthHandler(CONSUMER_KEY, CONSUMER_SECRET, CALLBACK_URL)
    try:
        # 連携アプリ認証用の URL を取得
        redirect_url = auth.get_authorization_url()
        # 認証後に必要な request_token を セッション に保存
        session['request_token'] = auth.request_token
    except tweepy.TweepError:
        return render_template("oauth_error.html", error = 0) #認証時エラーページ
    
    logging.info("---- API認証終了 ----")
    app.config["AUTH"] = auth #OAuthHandlerの保存
    return redirect(redirect_url) # redirect_urlのURLにリダイレクトする

#-------- テンプレートファイルダウンロードページ1 --------
@app.route('/download_1', methods=['GET']) #GET以外のリクエストを拒否
def download_1():
    filename = "Input_file.xlsm"
    #ファイルの出力
    #responseオブジェクトを作る
    response = make_response()   
    #ダウンロードデータをレスポンスオブジェクトのdataに設定 ここではファイルから読み込んだバイナリデータを設定
    response.data = open(os.path.join(app.config['DEFAULT_FOLDER'], filename), "rb").read()
    #レスポンスヘッダは設定されないためContent-Disposition: attachmentヘッダを手動で設定
    #attachment:ファイルのダウンロードタブを表示 inline:Webページ上で表示
    response.headers['Content-Disposition'] = 'attachment; filename=' + filename
    #レスポンスオブジェクトのmimetypeにダウンロードファイルのmimetypeを設定し、作成したレスポンスオブジェクトを戻り値として返却
    response.mimetype = XLSX_MIMETYPE
    return response

#-------- テンプレートファイルダウンロードページ2 --------
@app.route('/download_2', methods=['GET']) #GET以外のリクエストを拒否
def download_2():
    filename = "Input_file2.xlsm"
    #ファイルの出力
    #responseオブジェクトを作る
    response = make_response()   
    #ダウンロードデータをレスポンスオブジェクトのdataに設定 ここではファイルから読み込んだバイナリデータを設定
    response.data = open(os.path.join(app.config['DEFAULT_FOLDER'], filename), "rb").read()
    #レスポンスヘッダは設定されないためContent-Disposition: attachmentヘッダを手動で設定
    #attachment:ファイルのダウンロードタブを表示 inline:Webページ上で表示
    response.headers['Content-Disposition'] = 'attachment; filename=' + filename
    #レスポンスオブジェクトのmimetypeにダウンロードファイルのmimetypeを設定し、作成したレスポンスオブジェクトを戻り値として返却
    response.mimetype = XLSX_MIMETYPE
    return response

#-------- 使い方のページ --------
@app.route('/how_to_use', methods=['GET']) #GET
def how_to_use():
    return render_template("how_to_use.html")

#----* 拡張子の確認 *----
def allwed_file(filename):
    # .があるかどうかのチェックと、拡張子の確認
    # OKなら１、だめなら0
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS
    #.がfilenameにある 且つ .以下の文字列がアップロード可能な拡張子である
    #rsplit("区切る文字", 区切る数):文字列を後ろから指定文字で区切る filename.png => ["filename", "png"]
    #lower():小文字にする

#----* API取得 *----
def get_api():
    # request_token と oauth_verifier のチェック
    token = session.pop('request_token', None)
    #session.pop:sessionオブジェクトからrequest_tokenに紐付けられた項目を削除してその値を返す
    #もし第一引数の項目がなければ第二引数の項目を返す
    #第一引数の項目がなく、第二引数が指定されていないとエラーとなるので第一引数がない場合は何も返さないようにNoneとしている
    verifier = request.args.get('oauth_verifier')
    #Callbacl_URL?oauth_token=XXX&oauth_verifier=YYY の中からoauth_verifierのURLパラメータだけ取り出す
    if token is None or verifier is None: #request_tokenまたはoauth_verifierがないとき
        return False # 未認証ならFalseを返す
    
    #OAuthHandlerインスタンスの作成
    auth = auth = tweepy.OAuthHandler(CONSUMER_KEY, CONSUMER_SECRET, CALLBACK_URL)

    # Access token, Access token secret を取得．
    auth.request_token = token
    try:
        auth.get_access_token(verifier)
    except tweepy.TweepError:
        print("Error")
        return {}
    
    # tweepy で Twitter API にアクセス
    api = tweepy.API(auth)
    #Twitter_APIの保存
    if api:
        app.config['API']=api 
    
    return api
#----* ファイルのアップロード *----
def upload():
    # リクエストがポストかどうかの判別
    if request.method == 'POST':
        #ファイルが存在して且つxlsx形式じゃなかったら全部Noneを返す
        # ファイルがなかった場合の処理
        if 'file' not in request.files:
            return None
             # データの取り出し
        file = request.files['file']
        # ファイル名がなかった時の処理
        if file.filename == '':
                return None 
        #ファイルの種類がxlsxでないとき
        elif file and not allwed_file(file.filename):
                return None
        # ファイルのチェック
        elif file and allwed_file(file.filename):
            # 危険な文字を削除（サニタイズ処理）
            filename = secure_filename(file.filename)
            logging.info(filename)
            # ファイルの保存
            file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
            #os.path.join:引数の要素を繋げてパスを作る
            return filename
        else:
            return None
        
            
#----* ファイルのダウンロード *----
def download(filename):
    #ファイルの出力
    #responseオブジェクトを作る
    response = make_response()   
    #ダウンロードデータをレスポンスオブジェクトのdataに設定 ここではファイルから読み込んだバイナリデータを設定
    response.data = open(os.path.join(app.config['UPLOAD_FOLDER'], filename), "rb").read()
    #レスポンスヘッダは設定されないためContent-Disposition: attachmentヘッダを手動で設定
    #attachment:ファイルのダウンロードタブを表示 inline:Webページ上で表示
    response.headers['Content-Disposition'] = 'attachment; filename=' + filename
    #レスポンスオブジェクトのmimetypeにダウンロードファイルのmimetypeを設定し、作成したレスポンスオブジェクトを戻り値として返却
    response.mimetype = XLSX_MIMETYPE
    return response

#----* パターンマッチング *----
def pattern_match(tweet, num_pattern, circle_pattern, No): #引数(ツイートリスト,　番号パターン, サークル名パターン,　サークル情報リスト)
    uniq_no = [] #被りチェック用
    for twe in tweet:
        circle_num1 = num_pattern.findall(twe.user_name)                           #取得ツイートのユーザー名からスペース番号を抽出
        circle_num2 = num_pattern.findall(twe.text)                           #取得ツイートからスペース番号を抽出
        circle_name = circle_name_check(circle_pattern, twe.text, twe.description)      #取得ツイートのテキストまたはプロフィールからサークル名を抽出
        
        if len(circle_num1) == 0:                                   #ユーザー名から検出されなかったとき
            if len(circle_num2) == 1:                               #取得ツイートから1つだけ検出
                if circle_num2[0].replace('-','') not in uniq_no:   #被りがなければ
                    uniq_no.append(circle_num2[0].replace('-',''))  #被りチェックリストに追加
                    #サークル情報を追加(スペース番号, ユーザー名, ユーザーID, サークル名, キャラクター名, 抽出元URL)
                    No.append(Circle(circle_num2[0].replace('-',''),
                               twe.user_name,
                               twe.user_screen_name,
                               circle_name,
                               twe.chara,
                               twe.source_url))
                    
            else:                                                   #取得ツイートから複数検出または検出されなかったとき
                continue                                            #抽出不可、次のツイートへ
                    
        elif len(circle_num1) == 1:                                 #ユーザー名から1つだけ検出
            if len(circle_num2) == 0:                               #取得ツイートから検出されなかったとき
                if circle_num1[0].replace('-','') not in uniq_no:   #被りがなければ
                    uniq_no.append(circle_num1[0].replace('-',''))  #被りチェックリストに追加
                    No.append(Circle(circle_num1[0].replace('-',''), 
                               twe.user_name,
                               twe.user_screen_name,
                               circle_name,
                               twe.chara,
                               twe.source_url))
                    
            else:                                                           #取得ツイートから複数検出または1つのみ検出
                for num1 in circle_num1:                                    #ユーザー名のスペース番号
                    for num2 in circle_num2:                                #ツイート本文のスペース番号
                        if(num1.replace('-','') == num2.replace('-','')):   #ユーザー名とツイート本文のスペース番号を照合する(照合時、ハイフンを削除する)
                            if num1.replace('-','') not in uniq_no:         #被りがなければ
                                uniq_no.append(num1.replace('-',''))        #被りチェックリストに追加
                                No.append(Circle(num1.replace('-',''),
                               twe.user_name,
                               twe.user_screen_name,
                               circle_name,
                               twe.chara,
                               twe.source_url))   #一致する場合、リストに追加                    
                                
        else:                                              #ユーザー名から複数検出
            if len(circle_num2) == 0:                      #取得ツイートから検出されなかったとき
                continue                                   #抽出不可、次のツイートへ
                        
            else:                                          #取得ツイートから複数検出または1つのみ検出
                for num1 in circle_num1:                                    #ユーザー名のスペース番号
                    for num2 in circle_num2:                                #ツイート本文のスペース番号
                        if(num1.replace('-','') == num2.replace('-','')):   #ユーザー名とツイート本文のスペース番号を照合する(照合時、ハイフンを削除する)
                            if num1.replace('-','') not in uniq_no:         #被りがなければ
                                uniq_no.append(num1.replace('-',''))        #被りチェックリストに追加
                                No.append(Circle(num1.replace('-',''),twe.user_name,
                               twe.user_screen_name,
                               circle_name,
                               twe.chara,
                               twe.source_url))   #一致する場合、リストに追加
    
    return No

#----* サークルの名前を検出する *----
def circle_name_check(pattern, text, profile): #引数(パターン、マッチング元(ツイート), マッチング元(プロフィール))
    circle_name = re.search(pattern, text) #ツイートからマッチング
    if circle_name: # マッチしたとき
        circle_name = re.sub("サークル名?[「【『：:]|[」】』\n]", "", circle_name.group())     # サークル名のみ抽出
    else: # ツイートからマッチしなかったとき
        circle_name = re.search(pattern, profile)  # プロフからマッチング
        if circle_name: # マッチしたとき
            circle_name = re.sub("サークル名?[「【『：:]|[」】』\n]", "", circle_name.group()) # サークル名のみ抽出
        else: # それでもマッチしなかったとき                                                     
            circle_name = "no name"
    
    return circle_name

#----* キャラクター一覧の取得 *----
def chara_set(ws2):
    chara_list = [] #キャラクター一覧格納用
    space_list = [] #サークルリスト格納用
    for i,chara in enumerate(list(ws2.rows)):
        if i == 0:                          #初めの"Charater"というラベルを除外
            continue
        
        else:
            ws2_cells = list(chara)
            #サークル情報の格納
            space_list.append(Circle(ws2_cells[0], 
                                     ws2_cells[1],
                                     ws2_cells[2],
                                     ws2_cells[3],
                                     ws2_cells[4],
                                     ws2_cells[5],))
        
            if ws2_cells[4] not in chara_list:
                chara_list.append(ws2_cells[4])  #リストからキャラクター一覧の取得
    
    #キャラクターごとに色を決定        
    chara_dict = {}
    for j,chara_name in enumerate(chara_list):
        chara_dict[chara_name] = color[i % len(color)]

    return chara_dict, space_list

#----* シートの書式設定 *----
def sheet_format(ws2):
    #セルに罫線をつける/列幅を調整する
    #罫線のフォーマット
    border = op.styles.borders.Border(top=op.styles.borders.Side(style='thin', color='000000'), 
                                      bottom=op.styles.borders.Side(style='thin', color='000000'), 
                                      left=op.styles.borders.Side(style='thin', color='000000'),
                                      right=op.styles.borders.Side(style='thin', color='000000'))                
    #シート2書式
    for ws2_col in ws2.columns:                         #列の読み込み
        max_length = 0                                  #列幅初期値(0)
                    
        for ws2_cell in ws2_col:                        #現在の列のセルの読み込み
            if ws2_cell.value == None:                  #セルの値がない(None)とき
                continue
            else:                                       #セルに値があるとき
                ws2_cell.border = border                #罫線の設定
                
            if len(str(ws2_cell.value)) > max_length:   #文字列の長さが現在の列幅より大きいとき
                max_length = len(str(ws2_cell.value))   #文字列の長さを列幅に置き換え
                
        adjusted_width = round(max_length*2)  #列幅の調整(2倍)
        ws2.column_dimensions[op.utils.get_column_letter(ws2_col[0].column)].width = adjusted_width   
        #列幅の変更 シート.column_dimensions[列番号(アルファベット)].width = 列幅
        #op.utils.get_column_letter(列番号)列番号をアルファベットに変換

#----* 色付け *----
def coloring(space_list, chara_dict, space_position,  ws1, ws2): #space_list:サークルリスト, chara_dict:キャラクターごとの色の対応リスト, space_position:スペース番号の行列番号対応, ws1,ws2:作業シート
    #サークルリストの色付け
    for i, rows in enumerate(ws2):
        if i == 0: #1行目
            continue
        fill = op.styles.PatternFill(patternType='solid', fgColor = chara_dict[rows.cell.velue]) #色付け用フォーマット
        for cell in rows:
            cell.fill = fill #行の該当するセルを全て色付け
    
    #マップの色付け            
    for circle in space_list:
        fill = op.styles.PatternFill(patternType='solid', fgColor = chara_dict[circle.chara]) #色付け用フォーマット
        
        circle_num = re.sub("[ab]", "", circle.num) # abの削除
        if "b" in circle_num:   #スペース番号に"b"が含まれているとき
            ws1.cell(row = space_position[circle.num][0], column = space_position[circle.num][1] + 1).fill = fill                   #色付け
            add_comment(ws1.cell(row = space_position[circle.num][0], column = space_position[circle.num][1] + 1), circle)          #コメント追加
            ws1.cell(row = space_position[circle.num][0], column = space_position[circle.num][1] + 1).hyperlink = circle.source_url #ハイパーリンク
            
        else:   #スペース番号に"a"が含まれているとき または 机番号の区別がないとき
            ws1.cell(row = space_position[circle.num][0], column = space_position[circle.num][1]).fill = fill                       #色付け
            add_comment(ws1.cell(row = space_position[circle.num][0], column = space_position[circle.num][1] + 1), circle)          #コメント追加
            ws1.cell(row = space_position[circle.num][0], column = space_position[circle.num][1]).hyperlink = circle.source_url     #ハイパーリンク

#----* コメント付与 *----
def add_comment(cell, circleinfo): 
    #cell:コメントを付けるセルオブジェクト circleinfo:情報元インスタンス 
    #マップにコメントでサークル情報の付与
    cell.comment = op.comments.Comment("Writer:" + circleinfo.user_name \
                                      + "\nTwitter:" + circleinfo.user_screen_name \
                                      + "\nCircle:" + circleinfo.circle_name \
                                      + "\nchara:" + circleinfo.chara, "")
                                      #op.comments.Comment("コメント", "コメント作成者")
                                      #今回は
                                      #Writter:(作家)
                                      #Twitter:(作家のTwitterID)
                                      #Circle:(サークル名)
                                      #Chara:(キャラクター名)
                                      #とした
    cell.comment.width = 500 #コメント幅(横)の設定
    cell.comment.height  = 100 #コメント幅(縦)の設定 


#----** アプリの実行 **----
if __name__ == "__main__":
    app.run(debug=True)